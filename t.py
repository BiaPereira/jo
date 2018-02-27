import os
import numpy as np
import matplotlib.pyplot as plt
import openpyxl
from lmfit.models import GaussianModel, ConstantModel
from scipy.stats import norm, gaussian_kde
from bisect import bisect
from matplotlib.backends.backend_pdf import PdfPages
import math


#Opens file
import pandas as pd

book = openpyxl.load_workbook('LiCl_Logbook.xlsx')

sheet = book.active

print('\n\n\n\n')

#get the values for a given column
#values = book['Energia'].values
#print(a3)
#lent = len(a3)
def nome (filename):
    file = open(filename)
    print ('newfile',file)
    fileR = file.read()
    spl = fileR.split(' ', -1)
    print (type(spl))
    print (len)
    print(spl)
    file.close()
    return spl



#----------------------------------------fazer---------------------------------------
def energia (nome):
    s = str(nome).split('\\n', -1)
    print(len(s))
    print('s', s)
    energia = []
    for j in range (1,len(s),2):
        m=s[j]
        print('m',m)
        tam = len(m)
        print('vet',m[2:tam-4])
        for r in range(3, 100, 1):
            loc = sheet.cell(row=r, column=25)
            if loc.value == m[2:tam - 4]:
                l = 0
                cel = sheet.cell(row=r, column=4)
                valor = cel.value
                energia.append(valor)
                l = l + 1
                print('energia', energia)

    print('energia2', energia)
    return energia
#filename2 = 'LiCl2'
#file2 = open(filename2)
#print ('newfile',file2)
#fileR2 = file2.read()
#sp2 = fileR2.split(' ', -1)
#print (type(sp2))
#len2 = len(sp2)
#print (len2)
#area12,area22,c12,c22,fwhm12,fwhm22 = [],[],[],[],[],[]
#k=0
#for r in range (11,len2,87):
#    print(sp2[r:len2])
#    area12.append(float(sp2[r]))
#    print(area12)
#    c12.append(float(sp2[r+12]))
#    print(c12)
#    fwhm12.append(float(sp2[r+25]))
#    print(fwhm12)
#    area22.append(float((sp2[r+46])))
#    print(area22)
#    c22.append(float((sp2[r + 58])))
#    print(c22)
#    fwhm22.append(float(sp2[r + 71]))
#    print(fwhm22)
#    k=k+1

Snome = input('Nome Sem Relação')
Cnome = input('Nome Com Relação')

SRnome, SRarea1, SRarea2, SRc1, SRc2, SRfwhm1, SRfwhm2 = [], [], [], [], [], [], []
i = 0

spl = nome(Snome)
print(spl)
len1 = len(spl)
for w in range(0, len1 - 1, 83):
    print(spl[w:len1])
    SRnome.append(spl[w])
    print('nome', SRnome)
    SRarea1.append(float(spl[w + 11]))
    print('area1', SRarea1)
    SRc1.append(float(spl[w + 23]))
    print('c1', SRc1)
    SRfwhm1.append(float(spl[w + 36]))
    print('fw1', SRfwhm1)
    SRarea2.append(float((spl[w + 52])))
    print('area2', SRarea2)
    SRc2.append(float((spl[w + 64])))
    print('c2', SRc2)
    SRfwhm2.append(float(spl[w + 77]))
    print('fw2', SRfwhm2)
SRenergia = energia(SRnome)

#-------------------------------------------------------------------------------------------------------------

CRnome, CRarea1, CRarea2, CRc1, CRc2, CRfwhm1, CRfwhm2 = [], [], [], [], [], [], []


spl2 = nome(Cnome)
len1 = len(spl2)
for w in range(0, len1 - 1, 88):
    print(spl2[w:len1])
    CRnome.append(spl2[w])
    print('nome', CRnome)
    CRarea1.append(float(spl2[w + 11]))
    print('area1', CRarea1)
    CRc1.append(float(spl2[w + 23]))
    print('c1', CRc1)
    CRfwhm1.append(float(spl2[w + 36]))
    print('fw1', CRfwhm1)
    CRarea2.append(float((spl2[w + 57]))) #52
    print('area2', CRarea2)
    CRc2.append(float((spl2[w + 69])))
    print('c2', SRc2)
    CRfwhm2.append(float(spl2[w + 82]))
    print('fw2', CRfwhm2)
CRenergia = energia(CRnome)


plt.subplot(1,3,1)
plt.plot(SRenergia, SRarea1, 'b+', color = 'green', label = "Sem Relação 1")
plt.plot(SRenergia, SRarea2, 'b+', color = 'blue', label = "Sem Relação 2")
plt.plot(CRenergia,CRarea1, 'b+',color = 'orange', label = "Com Relação 1")
plt.plot(CRenergia, CRarea2, 'b+',color = 'black', label = "Com Relação 2")
plt.ylabel('Area')
plt.xlabel('Energia (keV)')
plt.legend(bbox_to_anchor=(-0.04, 0.05), loc=1, borderaxespad=0., ncol=1, title="Legenda:", fancybox=True)

plt.subplot(1,3,2)
plt.plot(SRenergia,SRc1, 'b+', color = 'green', label = "Sem Relação 1" )
plt.plot(SRenergia, SRc2, 'b+', color = 'blue',label = "Sem Relação 2")
plt.plot(CRenergia, CRc1, 'b+',color = 'orange', label = "Com Relação 1")
plt.plot(CRenergia, CRc2, 'b+',color = 'black', label = "Com Relação 2")
plt.ylabel('Centroide')
plt.xlabel('Energia (keV)')


plt.subplot(1,3,3)
s1 = plt.plot(SRenergia,SRfwhm1,'b+', color = 'green',  label = "Sem Relação 1")
s2 = plt.plot(SRenergia, SRfwhm2, 'b+', color = 'blue',label = "Sem Relação 2")
h1 = plt.plot(CRenergia,CRfwhm1, 'b+',color = 'orange', label = "Com Relação 1")
h2 = plt.plot(CRenergia, CRfwhm2, 'b+',color = 'black', label = "Com Relação 2")
plt.ylabel('FWHM')
plt.xlabel('Energia (keV)')

plt.subplots_adjust(top=0.9, left=0.1, right=0.9, bottom=0.12)


#plt.legend((s1, s2, h1, h2),  top=0.9, left=0.1, right=0.9, bottom=0.12)
#plt.subplot(2,2,4)

#plt.legend( loc=3, bbox_to_anchor=(0,0),
          # ncol=2, shadow=True, title="Legenda", fancybox=True)

#plt.legend("Sem Relação 2")
#plt.legend("Com Relação 1")
#plt.legend("Com Relação 2")


plt.tight_layout()
plt.show()


#get a data frame with selected columns
#FORMAT = ['Col_1', 'Col_2', 'Col_3']
#excel_data_selected = excel_data[FORMAT]




#print(Str[13:20])
#new=Str.split(' ',-1)
# print('novo', new)
# print(new[87])
# let = len(new)
# print (let)
# for i in range(87, let, 37):
  #  print (new[i])

#sheet = book.sheet_by_name('sheet')
#print(sheet)
#tam = len(filename)
